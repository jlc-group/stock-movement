"""Flask backend: serves the frontend and a JSON API backed by PostgreSQL.

Read endpoints:
    GET  /                  -> frontend
    GET  /api/health        -> DB connectivity
    GET  /api/products      -> full dataset (compact shape)

Admin (write) endpoints — require Authorization: Bearer <token>:
    POST /api/admin/login   -> exchange PIN for a session token
    POST /api/products      -> create a new product
    POST /api/movements     -> record/accumulate a daily in/out movement
    PUT  /api/movements     -> edit (overwrite) an existing day's movement
"""
import os
import sys
import uuid
import io
import json
import re
import urllib.request
import urllib.error
import urllib.parse
from functools import wraps
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

import psycopg2
import psycopg2.extras
from flask import Flask, jsonify, request, send_from_directory, send_file

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config
from brand_map import classify_brand

# URL ของ Script-Ecom launcher (ใช้ในหน้า "ออนไลน์" ดึงยอดออนไลน์มาลงช่อง online)
SCRIPT_ECOM_URL = os.getenv("SCRIPT_ECOM_URL", "http://127.0.0.1:4321")
SCRIPT_ECOM_APP_DIR = os.getenv(
    "SCRIPT_ECOM_APP_DIR",
    r"D:\AI_WORKSPACE\AI_Project\Github\Script-Ecom\JLC App\app",
)
PRINT_PLATFORMS = ("shopee", "lazada", "tiktok")
PRINT_STATUS_PATH = config.PROCESSED_DIR / "online_print_status.json"
# สถานะปริ้นระดับ "ออเดอร์" (ผูก order_id ไม่ผูกชื่อไฟล์ — ไฟล์ถูกดึงทับได้ สถานะไม่หาย)
ORDER_PRINT_STATUS_PATH = config.PROCESSED_DIR / "online_order_print_status.json"
PDF_PAGE_COUNT_CACHE = {}

app = Flask(__name__, static_folder=None)

# In-memory set of valid admin session tokens (cleared on restart).
ADMIN_TOKENS = set()

VALID_CATEGORIES = {"FG", "BTA", "PM", "BOX", "OTHER"}

# ---- Channels -----------------------------------------------------------
# Single source of truth for the movement "channel" dimension (added in the
# 001_add_channel migration). The DB column is `channel TEXT NOT NULL DEFAULT
# 'mixed'` with UNIQUE(product_id, movement_date, channel). 'mixed' is the
# legacy lane that pre-migration rows landed in, so a request that omits a
# channel keeps behaving exactly as before by defaulting to 'mixed'.
# `dir` is advisory metadata for the frontend dropdown only (in/out/both).
CHANNELS = [
    {"key": "online",     "label": "ออนไลน์",          "dir": "out"},
    {"key": "offline",    "label": "ออฟไลน์/หน้าร้าน", "dir": "out"},
    {"key": "wholesale",  "label": "ขายส่ง",           "dir": "out"},
    {"key": "redemption", "label": "แลกของรางวัล",     "dir": "out"},
    {"key": "kol",        "label": "KOL",              "dir": "out"},
    {"key": "influencer", "label": "Influencer",       "dir": "out"},
    {"key": "return",     "label": "รับคืน",           "dir": "in"},
    {"key": "receive",    "label": "รับเข้า",          "dir": "in"},
    {"key": "adjust",     "label": "ปรับยอด",          "dir": "both"},
    {"key": "mixed",      "label": "รวม (เดิม)",       "dir": "both"},
]
VALID_CHANNELS = {c["key"] for c in CHANNELS}


def norm_channel(v):
    """Normalize a request channel value.

    Missing/blank -> 'mixed' (backward compatible with pre-migration callers).
    Unknown key -> ValueError, so call sites can return a 400.
    """
    s = str(v or "").strip().lower() or "mixed"
    if s not in VALID_CHANNELS:
        raise ValueError(f"channel ไม่ถูกต้อง: {s}")
    return s


def get_conn():
    return psycopg2.connect(**config.DB)


def ensure_premium_warehouse_table():
    """Create the premium_warehouse table if it doesn't exist yet.

    Holds the per-product ลำลูกกา/ซอย8 split for premium products; the page's
    closing balance = ลำลูกกา + ซอย8 (written back to products.closing_balance).
    """
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {config.DB_SCHEMA}.premium_warehouse (
                product_id INT PRIMARY KEY
                    REFERENCES {config.DB_SCHEMA}.products(id) ON DELETE CASCADE,
                lamlukka NUMERIC NOT NULL DEFAULT 0,
                soi8     NUMERIC NOT NULL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[warn] ensure_premium_warehouse_table failed: {e}")


ensure_premium_warehouse_table()


def ensure_campaigns_table():
    """Create the campaigns + campaign_forecasts tables if they don't exist.

    Stores marketing campaigns shown on the /campaign page (previously kept in
    the browser's localStorage only). One campaign has a date range, a channel,
    a display color and a list of per-product forecast quantities.
    """
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {config.DB_SCHEMA}.campaigns (
                id         TEXT PRIMARY KEY,
                name       TEXT NOT NULL,
                channel    TEXT NOT NULL DEFAULT 'other',
                start_date DATE NOT NULL,
                end_date   DATE NOT NULL,
                color      TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {config.DB_SCHEMA}.campaign_forecasts (
                id          SERIAL PRIMARY KEY,
                campaign_id TEXT NOT NULL
                    REFERENCES {config.DB_SCHEMA}.campaigns(id) ON DELETE CASCADE,
                code        TEXT NOT NULL,
                name        TEXT,
                qty         NUMERIC NOT NULL DEFAULT 0
            )
        """)
        cur.execute(f"""
            CREATE INDEX IF NOT EXISTS idx_campaign_forecasts_cid
            ON {config.DB_SCHEMA}.campaign_forecasts(campaign_id)
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[warn] ensure_campaigns_table failed: {e}")


ensure_campaigns_table()


def ensure_weight_lots_table():
    """Create weight_lots table and seed initial data if empty."""
    _SEED = [
        ('M2', '010/24', 100.29, '2024-01-25'), ('M2', '011/24', 100.23, '2024-01-25'),
        ('M2', '012/24',  99.22, '2024-01-26'), ('M2', '013/24',  99.97, '2024-01-26'),
        ('M2', '015/24',  99.88, '2024-01-29'), ('M2', '016/24',  63.45, '2024-01-29'),
        ('M2', '020/24', 100.39, '2024-03-19'), ('M2', '021/24', 100.03, '2024-03-19'),
        ('M2', '022/24', 100.37, '2024-03-20'), ('M2', '023/24', 100.14, '2024-03-20'),
        ('M2', '025/24', 100.52, '2024-03-21'), ('M2', '026/24',  99.87, '2024-03-22'),
        ('M2', '027/24', 100.17, '2024-03-22'), ('M2', '028/24',  99.95, '2024-03-23'),
        ('M2', '029/24', 100.15, '2024-03-25'), ('M2', '030/24', 100.14, '2024-03-25'),
        ('M2', '031/24', 100.18, '2024-03-26'), ('M2', '032/24', 100.37, '2024-03-26'),
        ('M2', '033/24', 100.35, '2024-03-27'), ('M2', '034/24', 100.19, '2024-03-27'),
        ('M2', '035/24', 100.15, '2024-03-28'), ('M2', '036/24', 100.37, '2024-03-28'),
        ('M2', '037/24',  99.22, '2024-09-16'), ('M2', '038/24',  99.87, '2024-09-16'),
        ('M2', '039/24',  99.95, '2024-09-17'), ('M2', '040/24',  99.94, '2024-09-17'),
        ('M2', '041/24',  99.74, '2024-09-18'),
        ('K2', None,     929.56, None),
        ('L5', None,    1704.0,  None),
        ('M1', None,     437.0,  '2023-11-24'),
    ]
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS {config.DB_SCHEMA}.weight_lots (
                id            SERIAL PRIMARY KEY,
                product       TEXT NOT NULL,
                lot           TEXT,
                weight_kg     NUMERIC(10, 3) NOT NULL,
                produced_date DATE,
                created_at    TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute(f"SELECT COUNT(*) FROM {config.DB_SCHEMA}.weight_lots")
        if cur.fetchone()[0] == 0:
            for row in _SEED:
                cur.execute(f"""
                    INSERT INTO {config.DB_SCHEMA}.weight_lots
                        (product, lot, weight_kg, produced_date)
                    VALUES (%s, %s, %s, %s)
                """, row)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[warn] ensure_weight_lots_table failed: {e}")


ensure_weight_lots_table()


def num(x):
    """Decimal/None -> JSON-friendly int or float."""
    if x is None:
        return 0
    f = float(x)
    return int(f) if f.is_integer() else f


def script_ecom_print_dir(date_iso, platform):
    try:
        day = date.fromisoformat(str(date_iso or "").strip()).strftime("%d-%m-%Y")
    except ValueError:
        raise ValueError("invalid_date")
    platform = str(platform or "").strip().lower()
    if platform not in PRINT_PLATFORMS:
        raise ValueError("invalid_platform")
    return Path(SCRIPT_ECOM_APP_DIR) / "tmp" / day / platform / "print"


def parse_print_filename(filename, platform, folder_date):
    stem = Path(filename).stem
    suffix = f"_{folder_date}_{platform}"
    base = stem
    if suffix in base:
        base = base.split(suffix, 1)[0]
    parts = base.split("_", 1)
    carrier = parts[0] if len(parts) == 2 else ""
    sku_part = parts[1] if len(parts) == 2 else base
    qty = 1
    qty_match = re.search(r"_x(\d+)$", sku_part, flags=re.IGNORECASE)
    if qty_match:
        qty = int(qty_match.group(1))
        sku = sku_part[:qty_match.start()]
    else:
        sku = sku_part
    category = "OTHER"
    upper = sku.upper()
    if upper.startswith("SET_"):
        category = "SET"
    elif upper.startswith("DUO_"):
        category = "DUO"
    elif upper.startswith("MIX"):
        category = "MIX"
    else:
        m = re.match(r"([A-Z]+)", upper)
        if m:
            category = m.group(1)
    return {"carrier": carrier, "sku": sku, "pack_qty": qty, "category": category}


def pdf_page_count(path):
    try:
        stat = path.stat()
        key = (str(path), stat.st_mtime, stat.st_size)
        if key in PDF_PAGE_COUNT_CACHE:
            return PDF_PAGE_COUNT_CACHE[key]
    except Exception:
        key = None
    try:
        from pypdf import PdfReader
        count = len(PdfReader(str(path)).pages)
    except Exception:
        count = 1
    if key:
        PDF_PAGE_COUNT_CACHE[key] = count
    return count


def print_status_key(date_iso, platform, filename):
    return f"{date_iso}|{platform}|{filename}"


def load_print_statuses():
    try:
        if PRINT_STATUS_PATH.exists():
            with open(PRINT_STATUS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
    except Exception:
        return {}
    return {}


def save_print_statuses(statuses):
    config.PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    tmp = PRINT_STATUS_PATH.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(statuses, f, ensure_ascii=False, indent=2, sort_keys=True)
    os.replace(tmp, PRINT_STATUS_PATH)


def load_order_print_statuses():
    try:
        if ORDER_PRINT_STATUS_PATH.exists():
            with open(ORDER_PRINT_STATUS_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
    except Exception:
        return {}
    return {}


def save_order_print_statuses(statuses):
    config.PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    tmp = ORDER_PRINT_STATUS_PATH.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(statuses, f, ensure_ascii=False, indent=2, sort_keys=True)
    os.replace(tmp, ORDER_PRINT_STATUS_PATH)


# ---- "ยอดยืนยันสะสมต่อวัน" (working layer · ไม่ลด) -----------------------
# สะสม order_id ที่เคยเห็นในแต่ละวัน (union) → จำนวนออเดอร์ยืนยันไม่ลดแม้ขนส่งรับไปแล้ว
# + high-water ของ ledger snapshot (ชิ้น base unit) สำหรับยอดตัดสะสม
def confirmed_path(date_iso):
    safe = re.sub(r"[^0-9-]", "", str(date_iso))
    return config.PROCESSED_DIR / f"online_confirmed_{safe}.json"


def load_confirmed(date_iso):
    p = confirmed_path(date_iso)
    try:
        if p.exists():
            with open(p, "r", encoding="utf-8") as f:
                d = json.load(f)
                if isinstance(d, dict):
                    return d
    except Exception:
        pass
    return {"date": date_iso, "orders": {}, "pieces_hw": {}, "updated_at": ""}


def save_confirmed(date_iso, data):
    config.PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    p = confirmed_path(date_iso)
    tmp = p.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, p)


def accumulate_confirmed(date_iso):
    """สะสม 'ออเดอร์ที่เคยเห็น' ของวัน (union · ไม่ลด) = ออเดอร์ปัจจุบัน + รอบที่ freeze ไว้.

    สำคัญ: ต้อง backfill จากรอบด้วย เพราะออเดอร์ที่ขนส่งรับไปแล้วจะหายจากไฟล์ดิบ
    แต่ยังถูกจับไว้ในรอบที่เคยกดก่อนหน้า.
    """
    cur = load_confirmed(date_iso)
    orders = cur.get("orders") or {}
    pieces_hw = cur.get("pieces_hw") or {}
    added = 0

    def merge_orders(order_list):
        nonlocal added
        for o in order_list or []:
            plat = str(o.get("platform") or "")
            oid = str(o.get("order_id") or "")
            # ข้าม id สังเคราะห์ Lazada (LAZADA-STOCK-COUNT-SNAPSHOT / LAZADA-DELTA-* จาก
            # snapshot หรือรอบเก่าที่ backfill) — นับเฉพาะ order จริง. Lazada order จริง
            # (เลขล้วน) จะถูกนับ; pieces ยังมาจาก stock_snapshot (high-water) เหมือนเดิม.
            if not oid or oid.startswith("LAZADA-"):
                continue
            key = plat + "|" + oid
            if key not in orders:
                orders[key] = {"platform": plat, "items": o.get("items") or {}}
                added += 1

    def merge_hw(snap):
        for plat, items in (snap or {}).items():
            hw = pieces_hw.get(plat) or {}
            for sku, qty in (items or {}).items():
                n = float(qty or 0)
                if n > float(hw.get(sku) or 0):
                    hw[sku] = n
            pieces_hw[plat] = hw

    # 1) ออเดอร์ปัจจุบัน (จากการดึงล่าสุด)
    payload = script_ecom_json(
        "/api/stock/current-orders?date=" + urllib.parse.quote(date_iso), timeout=30)
    merge_orders(payload.get("orders"))
    merge_hw(payload.get("stock_snapshot"))

    # 2) backfill จากรอบที่ freeze ไว้ (จับออเดอร์ก่อนหน้าที่อาจส่งไปแล้ว)
    try:
        bl = script_ecom_json("/api/stock/batches?date=" + urllib.parse.quote(date_iso), timeout=15)
        for b in bl.get("batches") or []:
            full = script_ecom_json(
                "/api/stock/batch?id=" + urllib.parse.quote(str(b.get("id") or "")), timeout=15)
            merge_orders(full.get("orders"))
            merge_hw(full.get("stock_snapshot"))
    except Exception:
        pass  # ยังไม่มีรอบก็ข้ามไป

    # 3) ดึง order_id ที่ "ยกเลิก" ของวัน (Shopee+TikTok) → หาที่อยู่ในยอดสะสมเรา (intersection)
    #    ไม่ลบทิ้ง — เก็บ count ไว้โชว์ "ยกเลิกระหว่างวัน" + คำนวณยอดสุทธิ. พังก็ degrade เงียบ.
    cancel = {"available": False, "count": 0, "pieces": 0, "by_platform": {}, "errors": None}
    try:
        cx = script_ecom_json("/api/stock/cancelled?date=" + urllib.parse.quote(date_iso), timeout=60)
        cancelled_keys = set()
        for plat in ("shopee", "tiktok", "lazada"):
            for oid in cx.get(plat) or []:
                cancelled_keys.add(plat + "|" + str(oid))
        in_union = [k for k in orders if k in cancelled_keys]
        by_plat, pcs = {}, 0
        for k in in_union:
            p = orders[k].get("platform") or "?"
            by_plat[p] = by_plat.get(p, 0) + 1
            pcs += sum(float(v or 0) for v in (orders[k].get("items") or {}).values())
        cancel = {"available": True, "count": len(in_union), "pieces": pcs,
                  "by_platform": by_plat, "errors": cx.get("errors")}
    except Exception as e:
        cancel["fetch_error"] = str(e)

    cur.update({"date": date_iso, "orders": orders, "pieces_hw": pieces_hw,
                "cancelled": cancel,
                "updated_at": datetime.now().isoformat(timespec="seconds")})
    save_confirmed(date_iso, cur)
    return cur, added


def confirmed_summary(cur):
    orders = cur.get("orders") or {}
    pieces_hw = cur.get("pieces_hw") or {}
    by_plat = {}
    for o in orders.values():
        plat = o.get("platform") or "?"
        by_plat[plat] = by_plat.get(plat, 0) + 1
    plat_pieces = {p: sum(float(v or 0) for v in items.values())
                   for p, items in pieces_hw.items()}
    c = cur.get("cancelled") or {}
    cancelled_count = int(c.get("count") or 0)
    return {
        "date": cur.get("date"),
        "order_count": len(orders),
        "total_pieces": sum(plat_pieces.values()),
        "platforms_orders": by_plat,
        "platforms_pieces": plat_pieces,
        "updated_at": cur.get("updated_at"),
        "cancelled_available": bool(c.get("available")),
        "cancelled_count": cancelled_count,
        "cancelled_pieces": num(c.get("pieces") or 0),
        "cancelled_by_platform": c.get("by_platform") or {},
        "net_order_count": len(orders) - cancelled_count,
    }


def script_ecom_json(path, method="GET", payload=None, timeout=30):
    url = SCRIPT_ECOM_URL.rstrip("/") + path
    data = None
    headers = {"Accept": "application/json"}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode("utf-8"))


# ---- Admin auth ---------------------------------------------------------
def require_admin(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        token = auth[7:] if auth.startswith("Bearer ") else ""
        if not token or token not in ADMIN_TOKENS:
            return jsonify(error="unauthorized", detail="ต้องเข้าสู่โหมด Admin ก่อน"), 401
        return fn(*args, **kwargs)
    return wrapper


# ---- Shared helpers -----------------------------------------------------
def fetch_product(cur, product_id):
    """Return one product in the compact frontend shape (with tx list)."""
    cur.execute(f"""
        SELECT p.id, p.sheet_name, p.code, p.name, p.category_code,
               p.opening_balance, p.total_in, p.total_out, p.closing_balance, p.brand,
               COALESCE(pw.lamlukka, 0), COALESCE(pw.soi8, 0)
        FROM {config.DB_SCHEMA}.products p
        LEFT JOIN {config.DB_SCHEMA}.premium_warehouse pw ON pw.product_id = p.id
        WHERE p.id = %s
    """, (product_id,))
    p = cur.fetchone()
    if not p:
        return None
    cur.execute(f"""
        SELECT movement_date, qty_in, qty_out, balance, doc_no, note, channel,
               qty_shopee, qty_lazada, qty_tiktok
        FROM {config.DB_SCHEMA}.stock_movements
        WHERE product_id = %s
        ORDER BY movement_date, id
    """, (product_id,))
    # channel is at index 6; per-platform online qty appended at 7..9 (additive) so the
    # existing 0..6 fields keep their positions for every export/create/POST/PUT consumer.
    tx = [[
        r[0].isoformat() if r[0] else "",
        num(r[1]), num(r[2]), num(r[3]),
        r[4] or "", r[5] or "", r[6] or "mixed",
        num(r[7]), num(r[8]), num(r[9]),
    ] for r in cur.fetchall()]
    return {
        "id": p[0], "sheet": p[1], "code": p[2], "name": p[3] or "", "category": p[4],
        "opening": num(p[5]), "total_in": num(p[6]), "total_out": num(p[7]),
        "closing": num(p[8]), "brand": p[9] or classify_brand(p[2], p[3], p[4]),
        "lamlukka": num(p[10]), "soi8": num(p[11]), "tx": tx,
    }


def recompute_product(cur, product_id):
    """Recalculate running balances + product aggregates from opening balance.

    Must be called after any insert/update of that product's movements,
    because `balance` is cumulative and every later row shifts.
    """
    cur.execute(f"""
        SELECT opening_balance FROM {config.DB_SCHEMA}.products WHERE id = %s
    """, (product_id,))
    row = cur.fetchone()
    if not row:
        return
    opening = float(row[0] or 0)

    # After the channel migration there can be MANY rows per movement_date
    # (one per channel). Running balance folds every channel of a day; the
    # loop below already sums all returned rows, so the only requirement is a
    # deterministic order. We order by (movement_date, id) — i.e. insertion
    # order within a day — to keep cumulative balances stable across reruns.
    cur.execute(f"""
        SELECT id, qty_in, qty_out
        FROM {config.DB_SCHEMA}.stock_movements
        WHERE product_id = %s
        ORDER BY movement_date, id
    """, (product_id,))
    rows = cur.fetchall()

    running = opening
    total_in = 0.0
    total_out = 0.0
    for mid, qin, qout in rows:
        qin = float(qin or 0)
        qout = float(qout or 0)
        brought_forward = running
        running += qin - qout
        total_in += qin
        total_out += qout
        cur.execute(f"""
            UPDATE {config.DB_SCHEMA}.stock_movements
            SET balance = %s, brought_forward = %s
            WHERE id = %s
        """, (running, brought_forward, mid))

    cur.execute(f"""
        UPDATE {config.DB_SCHEMA}.products
        SET total_in = %s, total_out = %s, closing_balance = %s,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (total_in, total_out, running, product_id))


def apply_premium_warehouse_delta(cur, product_id, qty_in, qty_out, in_dest, brand=""):
    """For warehouse-backed products: qty_out is deducted from ซอย8,
    qty_in is added to the chosen warehouse (`soi8` or `lamlukka`).
    closing_balance = ลำลูกกา + ซอย8 (both are real warehouses holding stock;
    qty_out draws down ซอย8 so closing decreases).
    Overrides the movement-derived closing set by recompute_product."""
    cur.execute(f"""
        SELECT lamlukka, soi8 FROM {config.DB_SCHEMA}.premium_warehouse
        WHERE product_id = %s
    """, (product_id,))
    row = cur.fetchone()
    lam = float(row[0]) if row else 0.0
    soi = float(row[1]) if row else 0.0

    soi -= qty_out
    if in_dest == "lamlukka":
        lam += qty_in
    else:
        soi += qty_in

    closing = lam + soi

    cur.execute(f"""
        INSERT INTO {config.DB_SCHEMA}.premium_warehouse (product_id, lamlukka, soi8, updated_at)
        VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (product_id) DO UPDATE SET
            lamlukka = EXCLUDED.lamlukka, soi8 = EXCLUDED.soi8,
            updated_at = CURRENT_TIMESTAMP
    """, (product_id, lam, soi))
    cur.execute(f"""
        UPDATE {config.DB_SCHEMA}.products
        SET closing_balance = %s, updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (closing, product_id))


# ---- Read endpoints -----------------------------------------------------
@app.route("/")
def index():
    return send_from_directory(config.FRONTEND_DIR, "index.html")


@app.route("/campaign")
def campaign():
    return send_from_directory(config.FRONTEND_DIR, "campaign.html")


@app.route("/api/health")
def health():
    try:
        conn = get_conn()
        conn.close()
        return jsonify(status="ok", database="connected",
                       admin_enabled=bool(config.ADMIN_PIN))
    except Exception as e:
        return jsonify(status="error", database="unreachable", detail=str(e)), 503


@app.route("/api/products")
def products():
    """Full dataset in the compact shape the frontend expects."""
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute(f"""
        SELECT p.id, p.sheet_name, p.code, p.name, p.category_code,
               p.opening_balance, p.total_in, p.total_out, p.closing_balance, p.brand,
               COALESCE(pw.lamlukka, 0) AS lamlukka, COALESCE(pw.soi8, 0) AS soi8
        FROM {config.DB_SCHEMA}.products p
        LEFT JOIN {config.DB_SCHEMA}.premium_warehouse pw ON pw.product_id = p.id
        ORDER BY p.id
    """)
    prod_rows = cur.fetchall()

    cur.execute(f"""
        SELECT product_id, movement_date, qty_in, qty_out, balance, doc_no, note, channel,
               qty_shopee, qty_lazada, qty_tiktok
        FROM {config.DB_SCHEMA}.stock_movements
        ORDER BY product_id, movement_date, id
    """)
    # channel at index 6 (mirrors fetch_product); per-platform online qty at 7..9 (additive).
    # ORDER BY adds `id` so same-day multi-channel rows have a stable order
    # matching recompute_product's running-balance order.
    tx_by_product = defaultdict(list)
    for r in cur.fetchall():
        tx_by_product[r[0]].append([
            r[1].isoformat() if r[1] else "",
            num(r[2]), num(r[3]), num(r[4]),
            r[5] or "", r[6] or "", r[7] or "mixed",
            num(r[8]), num(r[9]), num(r[10]),
        ])

    cur.close()
    conn.close()

    products_out = [{
        "id": p["id"],
        "sheet": p["sheet_name"],
        "code": p["code"],
        "name": p["name"] or "",
        "category": p["category_code"],
        "brand": p["brand"] or classify_brand(p["code"], p["name"], p["category_code"]),
        "opening": num(p["opening_balance"]),
        "total_in": num(p["total_in"]),
        "total_out": num(p["total_out"]),
        "closing": num(p["closing_balance"]),
        "lamlukka": num(p["lamlukka"]),
        "soi8": num(p["soi8"]),
        "tx": tx_by_product.get(p["id"], []),
    } for p in prod_rows]

    return jsonify(
        generated_at=datetime.now().isoformat(),
        product_count=len(products_out),
        products=products_out,
    )


@app.route("/api/channels")
def channels():
    """Canonical channel metadata for the frontend dropdown (static, no DB).

    Derived from the CHANNELS module constant so the allow-list used by the
    POST/PUT validators and the keys advertised here can never drift apart.
    """
    return jsonify(channels=CHANNELS)


@app.route("/api/products/export")
def export_product():
    """Stream one product's full history as a real .xlsx file."""
    code = str(request.args.get("code", "")).strip()
    if not code:
        return jsonify(error="bad_request", detail="ต้องระบุรหัสสินค้า"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT id FROM {config.DB_SCHEMA}.products WHERE code = %s", (code,))
        row = cur.fetchone()
        if not row:
            return jsonify(error="not_found", detail=f"ไม่พบรหัสสินค้า {code}"), 404
        product = fetch_product(cur, row[0])
    finally:
        conn.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="F3F4F6")
    right = Alignment(horizontal="right")

    ws["A1"] = f"{product['code']} · {product['name']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"หมวด: {product['category']}  ·  Sheet: {product['sheet']}  ·  {len(product['tx'])} วัน"
    ws["A4"] = "ยอดยกมา"; ws["B4"] = product["opening"]
    ws["A5"] = "รับเข้าทั้งหมด"; ws["B5"] = product["total_in"]
    ws["A6"] = "จ่ายออกทั้งหมด"; ws["B6"] = product["total_out"]
    ws["A7"] = "คงเหลือปัจจุบัน"; ws["B7"] = product["closing"]
    for r in range(4, 8):
        ws[f"A{r}"].font = bold
        ws[f"B{r}"].alignment = right

    hdr = ["วันที่", "เลขที่", "รับ", "จ่าย", "คงเหลือ", "หมายเหตุ"]
    hrow = 9
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=hrow, column=i, value=h)
        c.font = bold
        c.fill = head_fill
    for j, t in enumerate(product["tx"], start=hrow + 1):
        d = t[0]
        ws.cell(row=j, column=1, value=("/".join(reversed(d.split("-"))) if d else ""))
        ws.cell(row=j, column=2, value=t[4])
        ws.cell(row=j, column=3, value=t[1])
        ws.cell(row=j, column=4, value=t[2])
        ws.cell(row=j, column=5, value=t[3])
        ws.cell(row=j, column=6, value=t[5])

    widths = [14, 16, 12, 12, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"stock_{code}.xlsx".replace("/", "-")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@app.route("/api/movements/export", methods=["POST"])
def export_movements():
    """Stream a flat daily-movement log (.xlsx) for the given product codes
    within an optional date range. Used by the product-list page "Export"
    button so the file honours whatever filters the user has applied: the
    client sends the already-filtered codes plus the active date range.
    """
    data = request.get_json(silent=True) or {}
    codes = data.get("codes") or []
    codes = [str(c).strip() for c in codes if str(c).strip()]
    date_from = str(data.get("from", "")).strip()
    date_to = str(data.get("to", "")).strip()
    for d in (date_from, date_to):
        if d:
            try:
                date.fromisoformat(d)
            except ValueError:
                return jsonify(error="bad_request", detail="วันที่ต้องอยู่ในรูปแบบ YYYY-MM-DD"), 400
    if not codes:
        return jsonify(error="bad_request", detail="ไม่มีรายการสินค้าให้ส่งออก"), 400

    where = ["p.code = ANY(%s)", "(m.qty_in <> 0 OR m.qty_out <> 0)"]
    params = [codes]
    if date_from:
        where.append("m.movement_date >= %s"); params.append(date_from)
    if date_to:
        where.append("m.movement_date <= %s"); params.append(date_to)

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT m.movement_date, p.code, p.name, p.brand, p.category_code,
                   m.qty_in, m.qty_out, m.balance, m.doc_no, m.note
            FROM {config.DB_SCHEMA}.stock_movements m
            JOIN {config.DB_SCHEMA}.products p ON p.id = m.product_id
            WHERE {' AND '.join(where)}
            ORDER BY m.movement_date, p.code, m.id
        """, params)
        rows = cur.fetchall()
    finally:
        conn.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Movements"
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="2563EB")
    right = Alignment(horizontal="right")

    rng = (f"{'/'.join(reversed(date_from.split('-')))}" if date_from else "เริ่มต้น") + \
          " – " + (f"{'/'.join(reversed(date_to.split('-')))}" if date_to else "ปัจจุบัน")
    ws["A1"] = "รายงานการเคลื่อนไหวสินค้า (รับเข้า / จ่ายออก)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"ช่วงวันที่: {rng}  ·  {len(rows)} รายการ"

    hdr = ["วันที่", "รหัสสินค้า", "ชื่อสินค้า", "แบรนด์", "หมวด",
           "รับ", "ออก", "คงเหลือ", "เลขที่เอกสาร", "หมายเหตุ"]
    hrow = 4
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=hrow, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    total_in = total_out = 0.0
    j = hrow
    for r in rows:
        j += 1
        d = r[0].isoformat() if r[0] else ""
        ws.cell(row=j, column=1, value=("/".join(reversed(d.split("-"))) if d else ""))
        ws.cell(row=j, column=2, value=r[1])
        ws.cell(row=j, column=3, value=r[2] or "")
        ws.cell(row=j, column=4, value=r[3] or "")
        ws.cell(row=j, column=5, value=r[4] or "")
        ws.cell(row=j, column=6, value=num(r[5]))
        ws.cell(row=j, column=7, value=num(r[6]))
        ws.cell(row=j, column=8, value=num(r[7]))
        ws.cell(row=j, column=9, value=r[8] or "")
        ws.cell(row=j, column=10, value=r[9] or "")
        total_in += float(r[5] or 0)
        total_out += float(r[6] or 0)

    j += 1
    ws.cell(row=j, column=5, value="รวม").font = bold
    tin = ws.cell(row=j, column=6, value=num(total_in)); tin.font = bold; tin.alignment = right
    tout = ws.cell(row=j, column=7, value=num(total_out)); tout.font = bold; tout.alignment = right

    widths = [14, 16, 34, 16, 8, 10, 10, 12, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tag = (date_from or "all") + "_" + (date_to or "all")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"movements_{tag}.xlsx",
    )


@app.route("/api/report/compare-export", methods=["POST"])
def export_report_compare():
    """Stream the Report-page 1/3/6-month comparison as a real .xlsx.
    The client posts the already-computed table (periods + per-product cells)
    so the file matches the on-screen numbers exactly — no server recompute.
    """
    data = request.get_json(silent=True) or {}
    periods = [str(p).strip() for p in (data.get("periods") or []) if str(p).strip()]
    rows = data.get("rows") or []
    if not periods or not rows:
        return jsonify(error="bad_request", detail="ไม่มีข้อมูลให้ส่งออก"), 400

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    bold = Font(bold=True)
    white = Font(bold=True, color="FFFFFF")
    grp_fill = PatternFill("solid", fgColor="2563EB")
    sub_fill = PatternFill("solid", fgColor="EFF6FF")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")

    ws["A1"] = "เปรียบเทียบสถิติย้อนหลัง " + " / ".join(periods)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"สร้างเมื่อ {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                f"  ·  {len(rows)} สินค้า  ·  ออกรวม · เฉลี่ย/วัน · วันออกมากสุด")

    g, s = 4, 5  # group-header row, sub-header row
    ws.cell(row=g, column=1, value="รหัส"); ws.merge_cells(start_row=g, start_column=1, end_row=s, end_column=1)
    ws.cell(row=g, column=2, value="ชื่อสินค้า"); ws.merge_cells(start_row=g, start_column=2, end_row=s, end_column=2)
    for col in (1, 2):
        hc = ws.cell(row=g, column=col); hc.font = bold; hc.fill = sub_fill; hc.alignment = center
    for pi, plabel in enumerate(periods):
        c0 = 3 + pi * 3
        gc = ws.cell(row=g, column=c0, value=plabel)
        ws.merge_cells(start_row=g, start_column=c0, end_row=g, end_column=c0 + 2)
        gc.font = white; gc.fill = grp_fill; gc.alignment = center
        for k, sublabel in enumerate(["ออกรวม", "เฉลี่ย/วัน", "วันออกมากสุด"]):
            sc = ws.cell(row=s, column=c0 + k, value=sublabel)
            sc.font = bold; sc.fill = sub_fill

    j = s
    for r in rows:
        j += 1
        ws.cell(row=j, column=1, value=str(r.get("code", "")))
        ws.cell(row=j, column=2, value=str(r.get("name", "")))
        cells = r.get("cells") or []
        for pi in range(len(periods)):
            c0 = 3 + pi * 3
            cell = cells[pi] if pi < len(cells) else {}
            oc = ws.cell(row=j, column=c0, value=num(cell.get("out", 0))); oc.alignment = right
            ac = ws.cell(row=j, column=c0 + 1, value=num(cell.get("avg", 0))); ac.alignment = right
            ws.cell(row=j, column=c0 + 2, value=str(cell.get("peak", "") or ""))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34
    for pi in range(len(periods)):
        c0 = 3 + pi * 3
        ws.column_dimensions[get_column_letter(c0)].width = 11
        ws.column_dimensions[get_column_letter(c0 + 1)].width = 11
        ws.column_dimensions[get_column_letter(c0 + 2)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="report_compare.xlsx",
    )


# ---- Admin endpoints ----------------------------------------------------
@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    if not config.ADMIN_PIN:
        return jsonify(error="admin_disabled",
                       detail="ยังไม่ได้ตั้งค่า ADMIN_PIN ใน .env"), 403
    data = request.get_json(silent=True) or {}
    pin = str(data.get("pin", "")).strip()
    if pin and pin == str(config.ADMIN_PIN):
        token = uuid.uuid4().hex
        ADMIN_TOKENS.add(token)
        return jsonify(status="ok", token=token)
    return jsonify(error="invalid_pin", detail="PIN ไม่ถูกต้อง"), 401


@app.route("/api/admin/logout", methods=["POST"])
@require_admin
def admin_logout():
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    ADMIN_TOKENS.discard(token)
    return jsonify(status="ok")


@app.route("/api/products", methods=["POST"])
@require_admin
def create_product():
    data = request.get_json(silent=True) or {}
    code = str(data.get("code", "")).strip()
    name = str(data.get("name", "")).strip()
    category = str(data.get("category", "")).strip().upper()
    brand = str(data.get("brand", "")).strip()
    sheet_name = str(data.get("sheet_name", "")).strip() or code
    try:
        opening = float(data.get("opening_balance", 0) or 0)
    except (TypeError, ValueError):
        return jsonify(error="bad_request", detail="ยอดยกมาต้องเป็นตัวเลข"), 400

    if not code:
        return jsonify(error="bad_request", detail="ต้องระบุรหัสสินค้า"), 400
    if category not in VALID_CATEGORIES:
        return jsonify(error="bad_request",
                       detail=f"หมวดต้องเป็น {', '.join(sorted(VALID_CATEGORIES))}"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        # Duplicate guards
        cur.execute(f"SELECT 1 FROM {config.DB_SCHEMA}.products WHERE code = %s", (code,))
        if cur.fetchone():
            return jsonify(error="conflict", detail=f"มีรหัส {code} อยู่แล้ว"), 409
        cur.execute(f"SELECT 1 FROM {config.DB_SCHEMA}.products WHERE sheet_name = %s", (sheet_name,))
        if cur.fetchone():
            return jsonify(error="conflict", detail=f"มี sheet '{sheet_name}' อยู่แล้ว"), 409

        cur.execute(f"""
            INSERT INTO {config.DB_SCHEMA}.products
                (sheet_name, code, name, category_code, brand,
                 opening_balance, total_in, total_out, closing_balance)
            VALUES (%s, %s, %s, %s, %s, %s, 0, 0, %s)
            RETURNING id
        """, (sheet_name, code, name, category, brand or classify_brand(code, name, category), opening, opening))
        product_id = cur.fetchone()[0]
        product = fetch_product(cur, product_id)
        conn.commit()
        return jsonify(status="ok", product=product), 201
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/premium/warehouse", methods=["PUT"])
@require_admin
def set_premium_warehouse():
    """Upsert the ลำลูกกา/ซอย8 split for one premium product and set its
    closing_balance = ลำลูกกา + ซอย8 (the premium page's source of truth)."""
    data = request.get_json(silent=True) or {}
    pid = data.get("id")
    try:
        lam = float(data.get("lamlukka", 0) or 0)
        soi = float(data.get("soi8", 0) or 0)
    except (TypeError, ValueError):
        return jsonify(error="bad_request", detail="ลำลูกกา/ซอย8 ต้องเป็นตัวเลข"), 400
    if not pid:
        return jsonify(error="bad_request", detail="ต้องระบุ id ของสินค้า"), 400
    if lam < 0 or soi < 0:
        return jsonify(error="bad_request", detail="จำนวนต้องไม่ติดลบ"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT brand FROM {config.DB_SCHEMA}.products WHERE id = %s", (pid,))
        row = cur.fetchone()
        if not row:
            return jsonify(error="not_found", detail="ไม่พบสินค้า"), 404
        brand = row[0] or ""
        closing = lam + soi
        cur.execute(f"""
            INSERT INTO {config.DB_SCHEMA}.premium_warehouse
                (product_id, lamlukka, soi8, updated_at)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (product_id) DO UPDATE SET
                lamlukka = EXCLUDED.lamlukka,
                soi8     = EXCLUDED.soi8,
                updated_at = CURRENT_TIMESTAMP
        """, (pid, lam, soi))
        cur.execute(f"""
            UPDATE {config.DB_SCHEMA}.products
            SET closing_balance = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (closing, pid))
        product = fetch_product(cur, pid)
        conn.commit()
        return jsonify(status="ok", product=product)
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/premium/transfer", methods=["POST"])
@require_admin
def transfer_premium_warehouse():
    """Internal warehouse move (warehouse-backed brands only): shift `qty` from
    ลำลูกกา to ซอย8. ลำลูกกา decreases, ซอย8 increases, closing (lam+soi) is
    UNCHANGED. Does NOT touch the ledger (stock_movements) or total_in/out."""
    data = request.get_json(silent=True) or {}
    pid = data.get("id")
    try:
        qty = float(data.get("qty", 0) or 0)
    except (TypeError, ValueError):
        return jsonify(error="bad_request", detail="จำนวนต้องเป็นตัวเลข"), 400
    if not pid:
        return jsonify(error="bad_request", detail="ต้องระบุ id ของสินค้า"), 400
    if qty <= 0:
        return jsonify(error="bad_request", detail="จำนวนที่ย้ายต้องมากกว่า 0"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT brand FROM {config.DB_SCHEMA}.products WHERE id = %s", (pid,))
        row = cur.fetchone()
        if not row:
            return jsonify(error="not_found", detail="ไม่พบสินค้า"), 404
        brand = row[0] or ""
        if brand not in ("สินค้าพรีเมี่ยม", "Beauterry"):
            return jsonify(error="bad_request",
                           detail="ย้ายคลังได้เฉพาะสินค้าพรีเมี่ยม/Beauterry"), 400

        cur.execute(f"""
            SELECT lamlukka, soi8 FROM {config.DB_SCHEMA}.premium_warehouse
            WHERE product_id = %s
        """, (pid,))
        wh = cur.fetchone()
        lam = float(wh[0]) if wh else 0.0
        soi = float(wh[1]) if wh else 0.0

        if qty > lam:
            return jsonify(error="bad_request",
                           detail=f"ย้ายได้สูงสุด {lam:g} ชิ้น (เท่ายอดลำลูกกา)"), 400

        lam -= qty
        soi += qty
        # closing = lam + soi is unchanged (internal move); still upsert so the
        # split is persisted and products.closing stays in sync.
        cur.execute(f"""
            INSERT INTO {config.DB_SCHEMA}.premium_warehouse
                (product_id, lamlukka, soi8, updated_at)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (product_id) DO UPDATE SET
                lamlukka = EXCLUDED.lamlukka,
                soi8     = EXCLUDED.soi8,
                updated_at = CURRENT_TIMESTAMP
        """, (pid, lam, soi))
        cur.execute(f"""
            UPDATE {config.DB_SCHEMA}.products
            SET closing_balance = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (lam + soi, pid))
        product = fetch_product(cur, pid)
        conn.commit()
        return jsonify(status="ok", product=product)
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


# ---- Campaigns ----------------------------------------------------------
CAMPAIGN_CHANNELS = {"lazada", "tiktok", "shopee", "other"}


def fetch_campaign(cur, cid):
    """Return one campaign in the shape the /campaign frontend expects."""
    cur.execute(f"""
        SELECT id, name, channel, start_date, end_date, color
        FROM {config.DB_SCHEMA}.campaigns WHERE id = %s
    """, (cid,))
    row = cur.fetchone()
    if not row:
        return None
    cur.execute(f"""
        SELECT code, name, qty FROM {config.DB_SCHEMA}.campaign_forecasts
        WHERE campaign_id = %s ORDER BY id
    """, (cid,))
    fcs = [{"code": r["code"], "name": r["name"] or "", "qty": num(r["qty"])}
           for r in cur.fetchall()]
    return {
        "id": row["id"], "name": row["name"], "channel": row["channel"],
        "start": row["start_date"].isoformat(), "end": row["end_date"].isoformat(),
        "color": row["color"], "forecasts": fcs,
    }


def validate_campaign(data):
    """Normalize+validate a campaign payload in place. Return error str or None."""
    name = str(data.get("name") or "").strip()
    if not name:
        return "กรุณากรอกชื่อแคมเปญ"
    channel = str(data.get("channel") or "").strip().lower()
    if channel not in CAMPAIGN_CHANNELS:
        return "ช่องทางไม่ถูกต้อง"
    start = str(data.get("start") or "").strip()
    end = str(data.get("end") or start).strip() or start
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", start):
        return "วันเริ่มไม่ถูกต้อง"
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", end):
        return "วันสิ้นสุดไม่ถูกต้อง"
    if end < start:
        return "วันสิ้นสุดต้องไม่ก่อนวันเริ่ม"
    data["name"] = name
    data["channel"] = channel
    data["start"] = start
    data["end"] = end
    data["color"] = (str(data.get("color")).strip() or None) if data.get("color") else None
    return None


def save_campaign_forecasts(cur, cid, forecasts):
    """Replace all forecast rows for a campaign with the given list."""
    cur.execute(
        f"DELETE FROM {config.DB_SCHEMA}.campaign_forecasts WHERE campaign_id = %s",
        (cid,),
    )
    for f in forecasts or []:
        code = str(f.get("code") or "").strip()
        if not code:
            continue
        try:
            qty = float(f.get("qty") or 0)
        except (TypeError, ValueError):
            qty = 0
        cur.execute(f"""
            INSERT INTO {config.DB_SCHEMA}.campaign_forecasts (campaign_id, code, name, qty)
            VALUES (%s, %s, %s, %s)
        """, (cid, code, str(f.get("name") or ""), qty))


@app.route("/api/campaigns")
def list_campaigns():
    """All campaigns (public, read-only) in the frontend shape."""
    conn = get_conn()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(
            f"SELECT id FROM {config.DB_SCHEMA}.campaigns ORDER BY start_date, name"
        )
        ids = [r["id"] for r in cur.fetchall()]
        camps = [fetch_campaign(cur, cid) for cid in ids]
        return jsonify(campaigns=camps)
    except Exception as e:
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/campaigns", methods=["POST"])
@require_admin
def create_campaign():
    data = request.get_json(silent=True) or {}
    err = validate_campaign(data)
    if err:
        return jsonify(error="bad_request", detail=err), 400
    cid = "c" + uuid.uuid4().hex[:12]
    conn = get_conn()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(f"""
            INSERT INTO {config.DB_SCHEMA}.campaigns
                (id, name, channel, start_date, end_date, color)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (cid, data["name"], data["channel"], data["start"],
              data["end"], data["color"]))
        save_campaign_forecasts(cur, cid, data.get("forecasts"))
        camp = fetch_campaign(cur, cid)
        conn.commit()
        return jsonify(status="ok", campaign=camp)
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/campaigns", methods=["PUT"])
@require_admin
def update_campaign():
    data = request.get_json(silent=True) or {}
    cid = data.get("id")
    if not cid:
        return jsonify(error="bad_request", detail="ต้องระบุ id ของแคมเปญ"), 400
    err = validate_campaign(data)
    if err:
        return jsonify(error="bad_request", detail=err), 400
    conn = get_conn()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute(
            f"SELECT 1 FROM {config.DB_SCHEMA}.campaigns WHERE id = %s", (cid,)
        )
        if not cur.fetchone():
            return jsonify(error="not_found", detail="ไม่พบแคมเปญ"), 404
        cur.execute(f"""
            UPDATE {config.DB_SCHEMA}.campaigns SET
                name = %s, channel = %s, start_date = %s, end_date = %s,
                color = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (data["name"], data["channel"], data["start"], data["end"],
              data["color"], cid))
        save_campaign_forecasts(cur, cid, data.get("forecasts"))
        camp = fetch_campaign(cur, cid)
        conn.commit()
        return jsonify(status="ok", campaign=camp)
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/campaigns", methods=["DELETE"])
@require_admin
def delete_campaign():
    data = request.get_json(silent=True) or {}
    cid = data.get("id") or request.args.get("id")
    if not cid:
        return jsonify(error="bad_request", detail="ต้องระบุ id ของแคมเปญ"), 400
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(
            f"DELETE FROM {config.DB_SCHEMA}.campaigns WHERE id = %s", (cid,)
        )
        conn.commit()
        return jsonify(status="ok")
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/weight-lots", methods=["GET"])
def get_weight_lots():
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT id, product, lot, weight_kg, produced_date, created_at
            FROM {config.DB_SCHEMA}.weight_lots
            ORDER BY product, produced_date NULLS LAST, lot NULLS LAST, id
        """)
        rows = cur.fetchall()
        return jsonify(lots=[{
            'id': r[0], 'product': r[1], 'lot': r[2],
            'weight_kg': float(r[3]) if r[3] is not None else None,
            'produced_date': r[4].isoformat() if r[4] else None,
        } for r in rows])
    finally:
        conn.close()


@app.route("/api/weight-lots", methods=["POST"])
@require_admin
def create_weight_lot():
    data = request.get_json(silent=True) or {}
    product = str(data.get('product') or '').strip()
    lot_raw = data.get('lot')
    lot = str(lot_raw).strip() if lot_raw else None
    pd_raw = data.get('produced_date')
    produced_date = str(pd_raw).strip() if pd_raw else None
    if not product:
        return jsonify(error="bad_request", detail="ต้องระบุ Product"), 400
    try:
        weight_kg = float(data.get('weight_kg', 0) or 0)
    except (TypeError, ValueError):
        return jsonify(error="bad_request", detail="น้ำหนักต้องเป็นตัวเลข"), 400
    if weight_kg <= 0:
        return jsonify(error="bad_request", detail="น้ำหนักต้องมากกว่า 0"), 400
    if produced_date:
        try:
            date.fromisoformat(produced_date)
        except ValueError:
            return jsonify(error="bad_request", detail="วันที่ไม่ถูกต้อง"), 400
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"""
            INSERT INTO {config.DB_SCHEMA}.weight_lots (product, lot, weight_kg, produced_date)
            VALUES (%s, %s, %s, %s)
            RETURNING id, product, lot, weight_kg, produced_date
        """, (product, lot, weight_kg, produced_date or None))
        r = cur.fetchone()
        conn.commit()
        return jsonify(status='ok', lot={
            'id': r[0], 'product': r[1], 'lot': r[2],
            'weight_kg': float(r[3]), 'produced_date': r[4].isoformat() if r[4] else None,
        }), 201
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/weight-lots/<int:lot_id>", methods=["DELETE"])
@require_admin
def delete_weight_lot(lot_id):
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM {config.DB_SCHEMA}.weight_lots WHERE id = %s", (lot_id,))
        if cur.rowcount == 0:
            conn.rollback()
            return jsonify(error="not_found"), 404
        conn.commit()
        return jsonify(status='ok')
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/movements", methods=["POST"])
@require_admin
def add_movement():
    data = request.get_json(silent=True) or {}
    code = str(data.get("code", "")).strip()
    pid_in = data.get("id")
    mv_date = str(data.get("date", "")).strip()
    doc_no = (str(data.get("doc_no", "")).strip() or None)
    note = (str(data.get("note", "")).strip() or None)
    in_dest = str(data.get("in_dest", "soi8")).strip() or "soi8"
    try:
        qty_in = float(data.get("qty_in", 0) or 0)
        qty_out = float(data.get("qty_out", 0) or 0)
    except (TypeError, ValueError):
        return jsonify(error="bad_request", detail="จำนวนต้องเป็นตัวเลข"), 400
    try:
        channel = norm_channel(data.get("channel"))  # missing/blank -> 'mixed'
    except ValueError:
        return jsonify(error="bad_request", detail="channel ไม่ถูกต้อง"), 400

    if not code and not pid_in:
        return jsonify(error="bad_request", detail="ต้องระบุรหัสสินค้า"), 400
    if in_dest not in ("soi8", "lamlukka"):
        return jsonify(error="bad_request", detail="คลังปลายทางต้องเป็น soi8 หรือ lamlukka"), 400
    if not mv_date:
        return jsonify(error="bad_request", detail="ต้องระบุวันที่"), 400
    try:
        date.fromisoformat(mv_date)
    except ValueError:
        return jsonify(error="bad_request", detail="วันที่ต้องอยู่ในรูปแบบ YYYY-MM-DD"), 400
    if qty_in < 0 or qty_out < 0:
        return jsonify(error="bad_request", detail="จำนวนห้ามติดลบ"), 400
    if qty_in == 0 and qty_out == 0:
        return jsonify(error="bad_request", detail="ต้องกรอกรับเข้าหรือจ่ายออกอย่างน้อย 1 ช่อง"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        # Prefer lookup by id (premium codes like UMBRELLA are not unique).
        if pid_in:
            cur.execute(f"SELECT id, brand FROM {config.DB_SCHEMA}.products WHERE id = %s", (pid_in,))
        else:
            cur.execute(f"SELECT id, brand FROM {config.DB_SCHEMA}.products WHERE code = %s", (code,))
        row = cur.fetchone()
        if not row:
            return jsonify(error="not_found", detail=f"ไม่พบรหัสสินค้า {code or pid_in}"), 404
        product_id, brand = row[0], row[1]
        # Warehouse-backed brands (ลำลูกกา/ซอย8 split, closing = lam + soi8).
        is_premium = brand in ("สินค้าพรีเมี่ยม", "Beauterry")

        # Same (day, channel) -> accumulate; otherwise a new row. Conflict
        # target is the post-migration key (product_id, movement_date, channel)
        # so each channel keeps its own lane for the same day.
        cur.execute(f"""
            INSERT INTO {config.DB_SCHEMA}.stock_movements
                (product_id, movement_date, doc_no, qty_in, qty_out, balance, note, channel)
            VALUES (%s, %s, %s, %s, %s, 0, %s, %s)
            ON CONFLICT (product_id, movement_date, channel) DO UPDATE SET
                qty_in  = {config.DB_SCHEMA}.stock_movements.qty_in  + EXCLUDED.qty_in,
                qty_out = {config.DB_SCHEMA}.stock_movements.qty_out + EXCLUDED.qty_out,
                doc_no  = COALESCE(EXCLUDED.doc_no, {config.DB_SCHEMA}.stock_movements.doc_no),
                note    = COALESCE(EXCLUDED.note,  {config.DB_SCHEMA}.stock_movements.note)
        """, (product_id, mv_date, doc_no, qty_in, qty_out, note, channel))

        recompute_product(cur, product_id)
        # Premium products are warehouse-backed: qty_out always draws from ซอย8,
        # qty_in lands in the chosen warehouse, and closing = ลำลูกกา + ซอย8
        # (overrides the movement-derived closing set by recompute_product).
        if is_premium:
            apply_premium_warehouse_delta(cur, product_id, qty_in, qty_out, in_dest, brand)
        product = fetch_product(cur, product_id)
        conn.commit()
        return jsonify(status="ok", product=product)
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/movements", methods=["PUT"])
@require_admin
def edit_movement():
    """Edit an existing daily movement: set (overwrite) qty/doc/note for a date,
    then recompute the running balance for the whole product."""
    data = request.get_json(silent=True) or {}
    code = str(data.get("code", "")).strip()
    mv_date = str(data.get("date", "")).strip()
    doc_no = (str(data.get("doc_no", "")).strip() or None)
    note = (str(data.get("note", "")).strip() or None)
    try:
        qty_in = float(data.get("qty_in", 0) or 0)
        qty_out = float(data.get("qty_out", 0) or 0)
    except (TypeError, ValueError):
        return jsonify(error="bad_request", detail="จำนวนต้องเป็นตัวเลข"), 400
    try:
        channel = norm_channel(data.get("channel"))  # missing/blank -> 'mixed'
    except ValueError:
        return jsonify(error="bad_request", detail="channel ไม่ถูกต้อง"), 400

    if not code:
        return jsonify(error="bad_request", detail="ต้องระบุรหัสสินค้า"), 400
    if not mv_date:
        return jsonify(error="bad_request", detail="ต้องระบุวันที่"), 400
    try:
        date.fromisoformat(mv_date)
    except ValueError:
        return jsonify(error="bad_request", detail="วันที่ต้องอยู่ในรูปแบบ YYYY-MM-DD"), 400
    if qty_in < 0 or qty_out < 0:
        return jsonify(error="bad_request", detail="จำนวนห้ามติดลบ"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT id, brand FROM {config.DB_SCHEMA}.products WHERE code = %s", (code,))
        row = cur.fetchone()
        if not row:
            return jsonify(error="not_found", detail=f"ไม่พบรหัสสินค้า {code}"), 404
        product_id, brand = row[0], row[1]

        # Target one (day, channel) lane — without the channel predicate this
        # would hit every channel row of the day after the migration.
        cur.execute(f"""
            UPDATE {config.DB_SCHEMA}.stock_movements
            SET qty_in = %s, qty_out = %s, doc_no = %s, note = %s
            WHERE product_id = %s AND movement_date = %s AND channel = %s
        """, (qty_in, qty_out, doc_no, note, product_id, mv_date, channel))
        if cur.rowcount == 0:
            conn.rollback()
            return jsonify(error="not_found",
                           detail=f"ไม่พบรายการของวันที่ {mv_date} ช่อง {channel}"), 404

        recompute_product(cur, product_id)
        # For warehouse-backed products recompute_product overwrites closing with
        # the movement-derived value; restore the warehouse-backed closing.
        # closing = ลำลูกกา + ซอย8 (both premium and Beauterry).
        if brand in ("สินค้าพรีเมี่ยม", "Beauterry"):
            cur.execute(f"""
                UPDATE {config.DB_SCHEMA}.products p
                SET closing_balance = pw.lamlukka + pw.soi8,
                    updated_at = CURRENT_TIMESTAMP
                FROM {config.DB_SCHEMA}.premium_warehouse pw
                WHERE p.id = %s AND pw.product_id = p.id
            """, (product_id,))
        product = fetch_product(cur, product_id)
        conn.commit()
        return jsonify(status="ok", product=product)
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


# ---- Bulk import (Excel) -----------------------------------------------
IMPORT_HEADERS = ["รหัสสินค้า", "วันที่", "รับเข้า", "จ่ายออก", "เลขที่เอกสาร", "หมายเหตุ"]


def _parse_import_date(v):
    """Excel cell -> 'YYYY-MM-DD' (or None if blank/invalid)."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return False  # present but unparseable


def _parse_import_num(v):
    """Excel cell -> float (0 for blank). Raises ValueError if non-numeric."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 0.0
    return float(str(v).replace(",", "").strip())


@app.route("/api/movements/import-template")
def import_template():
    """Download a blank .xlsx template with the expected header row."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "movements"
    head_fill = PatternFill("solid", fgColor="2563EB")
    for i, h in enumerate(IMPORT_HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    # Example rows (illustrative only).
    ws.append(["JH703-40G", "05/06/2026", 100, 0, "PO-001", "รับเข้าตัวอย่าง"])
    ws.append(["JH703-40G", "06/06/2026", 0, 30, "SO-001", "จ่ายออกตัวอย่าง"])
    for i, w in enumerate([16, 14, 10, 10, 16, 30], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="movement_import_template.xlsx",
    )


@app.route("/api/movements/import", methods=["POST"])
@require_admin
def import_movements():
    """Bulk-import daily in/out movements from an uploaded .xlsx.

    Single sheet, header row 1: รหัสสินค้า | วันที่ | รับเข้า | จ่ายออก |
    เลขที่เอกสาร | หมายเหตุ. Same-day rows accumulate (like POST /api/movements).
    Writes NOTHING unless every row is valid AND every code already exists —
    so the client can create missing products then re-upload safely.
    """
    from openpyxl import load_workbook

    # commit=1 actually writes; otherwise it's a dry-run preview (no writes).
    commit = str(request.form.get("commit", "")).strip() == "1"

    f = request.files.get("file")
    if not f:
        return jsonify(error="bad_request", detail="ไม่พบไฟล์ที่อัปโหลด"), 400
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True, read_only=True)
    except Exception:
        return jsonify(error="bad_request", detail="อ่านไฟล์ Excel ไม่สำเร็จ (.xlsx เท่านั้น)"), 400
    ws = wb.active

    # Map columns by header text (flexible to column order).
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def find(*keys):
        for idx, h in enumerate(header):
            hl = h.lower()
            if any(k in hl for k in keys):
                return idx
        return -1
    ci_code = find("รหัส", "code")
    ci_date = find("วันที่", "date")
    ci_in   = find("รับ", "in")
    ci_out  = find("จ่าย", "ออก", "out")
    ci_doc  = find("เอกสาร", "เลขที่", "doc")
    ci_note = find("หมายเหตุ", "note")
    if ci_code < 0 or ci_date < 0:
        return jsonify(error="bad_request",
                       detail="ไม่พบคอลัมน์ 'รหัสสินค้า' หรือ 'วันที่' ในไฟล์"), 400

    def cell(row, idx):
        return row[idx] if 0 <= idx < len(row) else None

    rows, errors = [], []
    for rno, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue  # skip blank line
        code = cell(row, ci_code)
        code = str(code).strip() if code is not None else ""
        d = _parse_import_date(cell(row, ci_date))
        try:
            qin = _parse_import_num(cell(row, ci_in))
            qout = _parse_import_num(cell(row, ci_out))
        except ValueError:
            errors.append({"row": rno, "detail": "จำนวนรับ/จ่ายต้องเป็นตัวเลข"})
            continue
        doc = cell(row, ci_doc); doc = str(doc).strip() if doc not in (None, "") else None
        note = cell(row, ci_note); note = str(note).strip() if note not in (None, "") else None

        if not code:
            errors.append({"row": rno, "detail": "ไม่มีรหัสสินค้า"}); continue
        if d is None:
            errors.append({"row": rno, "detail": "ไม่มีวันที่"}); continue
        if d is False:
            errors.append({"row": rno, "detail": "วันที่ไม่ถูกต้อง (ใช้ YYYY-MM-DD หรือ DD/MM/YYYY)"}); continue
        if qin < 0 or qout < 0:
            errors.append({"row": rno, "detail": "จำนวนห้ามติดลบ"}); continue
        if qin == 0 and qout == 0:
            errors.append({"row": rno, "detail": "ต้องมีรับเข้าหรือจ่ายออกอย่างน้อย 1 ช่อง"}); continue
        rows.append({"code": code, "date": d, "qty_in": qin, "qty_out": qout,
                     "doc_no": doc, "note": note})

    if errors:
        return jsonify(error="invalid_rows",
                       detail=f"พบข้อมูลผิดพลาด {len(errors)} แถว",
                       errors=errors[:50]), 400
    if not rows:
        return jsonify(error="empty", detail="ไม่พบข้อมูลในไฟล์"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT code, id, name, brand FROM {config.DB_SCHEMA}.products")
        code_to_id, code_to_name, code_to_brand = {}, {}, {}
        for c, i, nm, br in cur.fetchall():
            code_to_id[c] = i
            code_to_name[c] = nm or ""
            code_to_brand[c] = br or ""
        file_codes = {r["code"] for r in rows}
        unknown_set = {c for c in file_codes if c not in code_to_id}
        unknown = sorted(unknown_set)
        rows_per = defaultdict(int)
        for r in rows:
            if r["code"] in unknown_set:
                rows_per[r["code"]] += 1

        # Preview (dry-run): return the parsed rows for the user to review.
        if not commit:
            return jsonify(
                status="preview",
                total_rows=len(rows),
                unknown_codes=[{"code": c, "rows": rows_per[c]} for c in unknown],
                rows=[{
                    "code": r["code"],
                    "name": code_to_name.get(r["code"], ""),
                    "exists": r["code"] not in unknown_set,
                    "date": r["date"],
                    "qty_in": r["qty_in"],
                    "qty_out": r["qty_out"],
                    "doc_no": r["doc_no"] or "",
                    "note": r["note"] or "",
                } for r in rows],
            )

        if unknown:
            return jsonify(
                status="needs_products",
                detail=f"พบรหัสสินค้าที่ยังไม่มีในระบบ {len(unknown)} รายการ",
                unknown_codes=[{"code": c, "rows": rows_per[c]} for c in unknown],
                total_rows=len(rows),
            )

        affected = set()
        prem_delta = {}  # pid -> [qty_in_sum, qty_out_sum] for warehouse-backed brands
        for r in rows:
            pid = code_to_id[r["code"]]
            # Legacy Excel import has no per-row channel -> lands in 'mixed' so
            # totals match pre-migration behavior. Conflict target updated to
            # the new 3-col key (the old 2-col constraint no longer exists).
            cur.execute(f"""
                INSERT INTO {config.DB_SCHEMA}.stock_movements
                    (product_id, movement_date, doc_no, qty_in, qty_out, balance, note, channel)
                VALUES (%s, %s, %s, %s, %s, 0, %s, 'mixed')
                ON CONFLICT (product_id, movement_date, channel) DO UPDATE SET
                    qty_in  = {config.DB_SCHEMA}.stock_movements.qty_in  + EXCLUDED.qty_in,
                    qty_out = {config.DB_SCHEMA}.stock_movements.qty_out + EXCLUDED.qty_out,
                    doc_no  = COALESCE(EXCLUDED.doc_no, {config.DB_SCHEMA}.stock_movements.doc_no),
                    note    = COALESCE(EXCLUDED.note,  {config.DB_SCHEMA}.stock_movements.note)
            """, (pid, r["date"], r["doc_no"], r["qty_in"], r["qty_out"], r["note"]))
            affected.add(pid)
            # Warehouse-backed brands: remember the imported in/out so we can
            # draw it from ซอย8 after recompute (mirrors add_movement).
            if code_to_brand.get(r["code"]) in ("สินค้าพรีเมี่ยม", "Beauterry"):
                d = prem_delta.setdefault(pid, [0.0, 0.0])
                d[0] += r["qty_in"]
                d[1] += r["qty_out"]

        for pid in affected:
            recompute_product(cur, pid)
        # Warehouse-backed brands: qty_out draws from ซอย8, qty_in lands in ซอย8
        # (Excel import has no destination column), then closing = ลำลูกกา + ซอย8
        # (overrides the movement-derived closing set by recompute_product).
        for pid, (in_sum, out_sum) in prem_delta.items():
            apply_premium_warehouse_delta(cur, pid, in_sum, out_sum, "soi8")
        conn.commit()
        return jsonify(status="ok",
                       imported_rows=len(rows),
                       products_affected=len(affected))
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/online/import", methods=["POST"])
@require_admin
def online_import():
    """Idempotent SET-overwrite of the 'online' channel per (product, date).

    Body: { "date": "YYYY-MM-DD", "items": [ {"code": ..., "qty": ...}, ... ] }

    Designed for an automated feed (e.g. Script-Ecom): it OVERWRITES qty_out in
    the online lane (DO UPDATE SET qty_out = EXCLUDED.qty_out) instead of
    accumulating, so re-running the same day's import is safe. It only ever
    touches channel='online' — manually entered lanes (offline, redemption, …)
    are untouched. Unknown codes are skipped (never auto-created).
    """
    data = request.get_json(silent=True) or {}
    mv_date = str(data.get("date", "")).strip()
    items = data.get("items")

    if not mv_date:
        return jsonify(error="bad_request", detail="ต้องระบุวันที่"), 400
    try:
        date.fromisoformat(mv_date)
    except ValueError:
        return jsonify(error="bad_request", detail="วันที่ต้องอยู่ในรูปแบบ YYYY-MM-DD"), 400
    if not isinstance(items, list) or not items:
        return jsonify(error="bad_request", detail="ต้องมี items[] อย่างน้อย 1 รายการ"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT code, id FROM {config.DB_SCHEMA}.products")
        code_to_id = {c: i for c, i in cur.fetchall()}

        affected = set()
        written = 0
        skipped = 0
        for it in items:
            code = str((it or {}).get("code", "")).strip()
            try:
                qty = float((it or {}).get("qty", 0) or 0)
            except (TypeError, ValueError):
                return jsonify(error="bad_request",
                               detail=f"จำนวนของ {code or '(ไม่มีรหัส)'} ต้องเป็นตัวเลข"), 400
            if qty < 0:
                return jsonify(error="bad_request",
                               detail=f"จำนวนของ {code} ห้ามติดลบ"), 400

            pid = code_to_id.get(code)
            if pid is None:  # unknown code -> skip (do not auto-create)
                skipped += 1
                continue

            # SET-overwrite the online lane (contrast: add_movement accumulates).
            cur.execute(f"""
                INSERT INTO {config.DB_SCHEMA}.stock_movements
                    (product_id, movement_date, doc_no, qty_in, qty_out, balance, note, channel)
                VALUES (%s, %s, NULL, 0, %s, 0, NULL, 'online')
                ON CONFLICT (product_id, movement_date, channel) DO UPDATE SET
                    qty_out = EXCLUDED.qty_out
            """, (pid, mv_date, qty))
            affected.add(pid)
            written += 1

        for pid in affected:
            recompute_product(cur, pid)
        conn.commit()
        return jsonify(status="ok", written=written, skipped=skipped)
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/online/sync", methods=["POST"])
@require_admin
def online_sync():
    """หน้า 'ออนไลน์' กดปุ่มเดียว: ดึงยอดขายออนไลน์ล่าสุดจาก Script-Ecom แล้วลงช่อง 'online' ให้อัตโนมัติ.

    body (ไม่บังคับ): { "date": "YYYY-MM-DD" } — ไม่ส่ง = วันล่าสุดที่ Script-Ecom มี.
    เขียนเฉพาะช่อง online (SET-overwrite, ยิงซ้ำได้) — ไม่แตะช่องที่กรอกมือ.
    """
    data = request.get_json(silent=True) or {}
    mv_date = str(data.get("date", "")).strip()

    # 1) ดึง "ยอดตัดสะสมทั้งวัน" (รวมทุกรอบที่ freeze = ชิ้นเดี่ยว ไม่ลด) จาก Script-Ecom launcher
    #    ใช้ยอดสะสมแทนยอดสด (stock_count) เพราะของที่แพ็คส่งไปแล้วต้องคงตัดไว้ ไม่งั้นตัดสต็อกขาด
    url = SCRIPT_ECOM_URL.rstrip("/") + "/api/stock/propose-cumulative"
    if mv_date:
        url += "?date=" + mv_date
    try:
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as r:
            payload = json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as he:
        # Script-Ecom เชื่อมได้ แต่ตอบ error (ที่พบบ่อย: ยังไม่มีข้อมูลออเดอร์ของวันนั้น)
        reason = ""
        try:
            reason = (json.loads(he.read().decode("utf-8")) or {}).get("error", "")
        except Exception:
            reason = ""
        if not reason:
            reason = f"HTTP {he.code}"
        return jsonify(error="no_data",
                       detail=f"Script-Ecom: {reason} — ลองเลือกวันที่ที่ดึงออเดอร์ไว้แล้ว หรือกดดึงออเดอร์ของวันนั้นใน Script-Ecom ก่อน"), 424
    except Exception as e:
        return jsonify(error="upstream_unreachable",
                       detail=f"เชื่อม Script-Ecom ไม่ได้ ({url}) — เปิด start_ui.bat ของ Script-Ecom ค้างไว้ก่อน [{e}]"), 424

    p_date = (payload.get("date") or mv_date or "").strip()
    proposed = payload.get("proposed") or []
    try:
        date.fromisoformat(p_date)
    except ValueError:
        return jsonify(error="bad_request", detail=f"วันที่จาก Script-Ecom ไม่ถูกต้อง: {p_date}"), 400

    # 2) ลงช่อง online (idempotent) — เฉพาะรหัสที่มีใน ecom_stock
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT code, id FROM {config.DB_SCHEMA}.products")
        code_to_id = {c: i for c, i in cur.fetchall()}
        affected, applied, skipped = set(), [], []
        doc_no = "ONLINE-" + p_date.replace("-", "")
        for it in proposed:
            code = str((it or {}).get("code", "")).strip()
            try:
                qty = float((it or {}).get("qty", 0) or 0)
            except (TypeError, ValueError):
                qty = 0
            pid = code_to_id.get(code)
            if pid is None:
                skipped.append(code)
                continue
            sp = float((it or {}).get("shopee", 0) or 0)
            lz = float((it or {}).get("lazada", 0) or 0)
            tt = float((it or {}).get("tiktok", 0) or 0)
            cur.execute(f"""
                INSERT INTO {config.DB_SCHEMA}.stock_movements
                    (product_id, movement_date, doc_no, qty_in, qty_out, balance, note, channel,
                     qty_shopee, qty_lazada, qty_tiktok)
                VALUES (%s, %s, %s, 0, %s, 0, %s, 'online', %s, %s, %s)
                ON CONFLICT (product_id, movement_date, channel) DO UPDATE SET
                    qty_out = EXCLUDED.qty_out, doc_no = EXCLUDED.doc_no, note = EXCLUDED.note,
                    qty_shopee = EXCLUDED.qty_shopee, qty_lazada = EXCLUDED.qty_lazada,
                    qty_tiktok = EXCLUDED.qty_tiktok
            """, (pid, p_date, doc_no, qty, "ตัดสต็อกออนไลน์ (auto)", sp, lz, tt))
            affected.add(pid)
            applied.append({
                "code": code,
                "source_sku": (it or {}).get("sku", ""),
                "qty": qty,
                "name": (it or {}).get("name", ""),
                "shopee": (it or {}).get("shopee", 0) or 0,
                "lazada": (it or {}).get("lazada", 0) or 0,
                "tiktok": (it or {}).get("tiktok", 0) or 0,
            })
        for pid in affected:
            recompute_product(cur, pid)
        conn.commit()
        return jsonify(status="ok", date=p_date, written=len(applied),
                       skipped=len(skipped), skipped_codes=skipped, items=applied,
                       platforms=payload.get("platforms") or {})
    except Exception as e:
        conn.rollback()
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/online/preview", methods=["POST"])
@require_admin
def online_preview():
    """โหมดพิสูจน์ (read-only): ดึง 'ยอดออนไลน์ที่ควรตัด' จาก Script-Ecom มาแสดงเฉยๆ — ไม่เขียนลง DB เลย.
    ให้ admin เอาไปเทียบกับยอดที่กรอกมือ เพื่อพิสูจน์ความถูกต้องก่อนเปิดโหมดอัตโนมัติ."""
    data = request.get_json(silent=True) or {}
    mv_date = str(data.get("date", "")).strip()
    url = SCRIPT_ECOM_URL.rstrip("/") + "/api/stock/propose"
    if mv_date:
        url += "?date=" + mv_date
    try:
        req = urllib.request.Request(url, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=30) as r:
            payload = json.loads(r.read().decode("utf-8"))
    except urllib.error.HTTPError as he:
        reason = ""
        try:
            reason = (json.loads(he.read().decode("utf-8")) or {}).get("error", "")
        except Exception:
            reason = ""
        if not reason:
            reason = f"HTTP {he.code}"
        return jsonify(error="no_data",
                       detail=f"Script-Ecom: {reason} — ลองเลือกวันที่ที่ดึงออเดอร์ไว้แล้ว"), 424
    except Exception as e:
        return jsonify(error="upstream_unreachable",
                       detail=f"เชื่อม Script-Ecom ไม่ได้ ({url}) [{e}]"), 424

    p_date = (payload.get("date") or mv_date or "").strip()
    proposed = payload.get("proposed") or []
    try:
        date.fromisoformat(p_date)
    except ValueError:
        return jsonify(error="bad_request", detail=f"วันที่จาก Script-Ecom ไม่ถูกต้อง: {p_date}"), 400

    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT code, id, name FROM {config.DB_SCHEMA}.products")
        code_to = {c: {"id": i, "name": n} for c, i, n in cur.fetchall()}
        # ยอดช่อง online ที่ "เคยลงไว้แล้ว" ของวันนั้น (ถ้ามี) เพื่อเทียบ
        cur.execute(f"""
            SELECT product_id, qty_out FROM {config.DB_SCHEMA}.stock_movements
            WHERE movement_date = %s AND channel = 'online'
        """, (p_date,))
        existing = {pid: q for pid, q in cur.fetchall()}

        items, skipped, total = [], [], 0
        for it in proposed:
            code = str((it or {}).get("code", "")).strip()
            try:
                qty = float((it or {}).get("qty", 0) or 0)
            except (TypeError, ValueError):
                qty = 0
            info = code_to.get(code)
            if info is None:
                skipped.append({"code": code, "qty": qty, "name": (it or {}).get("name", "")})
                continue
            total += qty
            items.append({
                "code": code, "name": info["name"], "qty": qty,
                "source_sku": (it or {}).get("sku", ""),
                "shopee": (it or {}).get("shopee", 0) or 0,
                "lazada": (it or {}).get("lazada", 0) or 0,
                "tiktok": (it or {}).get("tiktok", 0) or 0,
                "already_online": existing.get(info["id"], 0),
            })
        items.sort(key=lambda x: x["code"])

        # ---- หักยอดที่ถูก "บันทึกรอบยืนยัน" ไปแล้ว → เหลือเฉพาะออเดอร์ใหม่ที่ยังไม่เข้ารอบ ----
        delta_applied = False
        batched_pcs = 0
        batches_count = 0
        latest_batch_label = ""
        try:
            bl = script_ecom_json("/api/stock/batches?date=" + urllib.parse.quote(p_date), timeout=15)
            blist = list(bl.get("batches") or [])
            batches_count = len(blist)
            if blist:
                blist.sort(key=lambda b: str(b.get("created_at") or ""))
                latest = blist[-1]
                latest_batch_label = str(latest.get("label") or "")
                full = script_ecom_json(
                    "/api/stock/batch?id=" + urllib.parse.quote(str(latest.get("id") or "")), timeout=15)
                snap = full.get("stock_snapshot") or {}
                if snap:  # batch รุ่นใหม่ (v2) เท่านั้น — รุ่นเก่าไม่มี snapshot ก็แสดงยอดเต็มตามเดิม
                    delta_applied = True
                    before = total
                    remaining = []
                    for it in items:
                        sku = str(it.get("source_sku") or "")
                        rem_total = 0
                        for plat in ("shopee", "lazada", "tiktok"):
                            have = float(it.get(plat) or 0)
                            used = float((snap.get(plat) or {}).get(sku) or 0)
                            rem = have - used
                            it[plat] = rem if rem > 0 else 0
                            rem_total += it[plat]
                        it["qty"] = rem_total
                        if rem_total > 0:
                            remaining.append(it)
                    items = remaining
                    total = sum(it["qty"] for it in items)
                    batched_pcs = max(0, before - total)
        except Exception:
            pass  # Script-Ecom ล่ม/ไม่มี batch → โชว์ยอดเต็มตามเดิม

        # ---- ยอด "สะสมทุกรอบ" = ตัวเดียวกับที่ปุ่มเขียนจริง (sync) จะตัด ----
        # ดึงมาโชว์ให้ผู้ใช้เห็นเลขที่จะเขียนจริง (ไม่ใช่เลข leftover) + กันสับสน/ตัดขาด
        cumulative_pcs = 0.0
        cumulative_items = 0
        cumulative_rounds = 0
        cumulative_skipped = []  # รหัสในยอดสะสมที่ map ไม่ได้ → จะไม่ถูกตัด (โชว์เตือนก่อนเขียน)
        try:
            cumj = script_ecom_json(
                "/api/stock/propose-cumulative?date=" + urllib.parse.quote(p_date), timeout=20)
            cumulative_rounds = int(cumj.get("batches_count") or 0)
            for ci in (cumj.get("proposed") or []):
                ccode = str((ci or {}).get("code", "")).strip()
                try:
                    cqty = float((ci or {}).get("qty", 0) or 0)
                except (TypeError, ValueError):
                    cqty = 0.0
                if ccode in code_to:  # นับเฉพาะรหัสที่ map ได้ (เหมือนตอนเขียนจริง)
                    cumulative_items += 1
                    cumulative_pcs += cqty
                else:  # map ไม่ได้ → จะไม่ถูกตัด เก็บไว้เตือนผู้ใช้
                    cumulative_skipped.append(
                        {"code": ccode, "qty": cqty, "name": (ci or {}).get("name", "")})
        except Exception:
            pass  # Script-Ecom ล่ม → ไม่โชว์ยอดสะสม (ปุ่มยังทำงานตามเดิม)

        return jsonify(status="ok", date=p_date, mode="preview",
                       total_qty=total, matched=len(items), skipped=len(skipped),
                       items=items, skipped_items=skipped,
                       delta_applied=delta_applied, batched_pcs=batched_pcs,
                       batches_count=batches_count, latest_batch_label=latest_batch_label,
                       cumulative_pcs=cumulative_pcs, cumulative_items=cumulative_items,
                       cumulative_rounds=cumulative_rounds,
                       cumulative_skipped=cumulative_skipped,
                       platforms=payload.get("platforms") or {})
    except Exception as e:
        return jsonify(error="server_error", detail=str(e)), 500
    finally:
        conn.close()


@app.route("/api/online/pull", methods=["POST"])
@require_admin
def online_pull():
    """สั่ง Script-Ecom ดึงออเดอร์ (ค่าเริ่มต้น set2_all = Shopee+Lazada+TikTok) — fire แล้วให้หน้าเว็บ poll สถานะ."""
    data = request.get_json(silent=True) or {}
    job = str(data.get("job", "set2_all")).strip() or "set2_all"
    url = SCRIPT_ECOM_URL.rstrip("/") + "/api/run"
    body = json.dumps({"job": job}).encode("utf-8")
    try:
        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=30) as r:
            payload = json.loads(r.read().decode("utf-8"))
        return jsonify(status="ok", started=True, job=job, launcher=payload)
    except urllib.error.HTTPError as he:
        try:
            reason = (json.loads(he.read().decode("utf-8")) or {}).get("error", f"HTTP {he.code}")
        except Exception:
            reason = f"HTTP {he.code}"
        return jsonify(error="pull_failed", detail=f"Script-Ecom: {reason}"), 424
    except Exception as e:
        return jsonify(error="upstream_unreachable",
                       detail=f"เชื่อม Script-Ecom ไม่ได้ — เปิด start_ui.bat ของ Script-Ecom ค้างไว้ก่อน [{e}]"), 424


@app.route("/api/online/pull/status")
@require_admin
def online_pull_status():
    """อ่านสถานะงานดึงของ Script-Ecom (proxy /api/state) ให้หน้าเว็บ poll."""
    url = SCRIPT_ECOM_URL.rstrip("/") + "/api/state"
    try:
        with urllib.request.urlopen(url, timeout=15) as r:
            payload = json.loads(r.read().decode("utf-8"))
        return jsonify(status="ok", job=payload.get("job", {}))
    except Exception as e:
        return jsonify(error="upstream_unreachable", detail=str(e)), 424


@app.route("/api/online/accumulate", methods=["POST"])
@require_admin
def online_accumulate():
    """สะสม 'ออเดอร์ที่ยืนยันแล้ว' ของวันนั้น (union order_id · ไม่ลด) — เรียกหลังดึงออเดอร์เสร็จ."""
    data = request.get_json(silent=True) or {}
    date_iso = str(data.get("date", "")).strip()
    if not date_iso:
        return jsonify(error="bad_request", detail="ต้องระบุวันที่"), 400
    try:
        cur, added = accumulate_confirmed(date_iso)
    except urllib.error.HTTPError as he:
        try:
            reason = (json.loads(he.read().decode("utf-8")) or {}).get("error", f"HTTP {he.code}")
        except Exception:
            reason = f"HTTP {he.code}"
        return jsonify(error="upstream", detail=f"Script-Ecom: {reason}"), 424
    except Exception as e:
        return jsonify(error="upstream_unreachable", detail=str(e)), 424
    return jsonify(status="ok", added=added, **confirmed_summary(cur))


@app.route("/api/online/confirmed")
@require_admin
def online_confirmed():
    """อ่านยอดยืนยันสะสมของวันนั้น (ไม่เรียก Script-Ecom — อ่านจากที่สะสมไว้)."""
    date_iso = str(request.args.get("date", "")).strip()
    if not date_iso:
        return jsonify(error="bad_request", detail="ต้องระบุวันที่"), 400
    return jsonify(status="ok", **confirmed_summary(load_confirmed(date_iso)))


@app.route("/api/online/batches")
@require_admin
def online_batches():
    date_iso = str(request.args.get("date", "")).strip()
    path = "/api/stock/batches"
    if date_iso:
        path += "?date=" + urllib.parse.quote(date_iso)
    try:
        return jsonify(script_ecom_json(path, timeout=15))
    except urllib.error.HTTPError as he:
        try:
            reason = (json.loads(he.read().decode("utf-8")) or {}).get("error", f"HTTP {he.code}")
        except Exception:
            reason = f"HTTP {he.code}"
        return jsonify(error="batch_failed", detail=f"Script-Ecom: {reason}"), 424
    except Exception as e:
        return jsonify(error="upstream_unreachable", detail=str(e)), 424


@app.route("/api/online/batches", methods=["POST"])
@require_admin
def online_create_batch():
    data = request.get_json(silent=True) or {}
    date_iso = str(data.get("date", "")).strip()
    label = str(data.get("label", "")).strip()
    note = str(data.get("note", "")).strip()
    try:
        payload = script_ecom_json(
            "/api/stock/batches",
            method="POST",
            payload={"date": date_iso, "label": label, "note": note},
            timeout=30,
        )
        return jsonify(payload)
    except urllib.error.HTTPError as he:
        try:
            reason = (json.loads(he.read().decode("utf-8")) or {}).get("error", f"HTTP {he.code}")
        except Exception:
            reason = f"HTTP {he.code}"
        return jsonify(error="batch_failed", detail=f"Script-Ecom: {reason}"), 424
    except Exception as e:
        return jsonify(error="upstream_unreachable", detail=str(e)), 424


@app.route("/api/online/batch")
@require_admin
def online_batch_detail():
    """รายละเอียดรอบยืนยัน 1 รอบ (proxy Script-Ecom) + merge สถานะปริ้นรายออเดอร์ของเรา."""
    bid = str(request.args.get("id", "")).strip()
    if not bid:
        return jsonify(error="bad_request", detail="ต้องระบุ id ของรอบ"), 400
    try:
        payload = script_ecom_json("/api/stock/batch?id=" + urllib.parse.quote(bid), timeout=30)
    except urllib.error.HTTPError as he:
        try:
            reason = (json.loads(he.read().decode("utf-8")) or {}).get("error", f"HTTP {he.code}")
        except Exception:
            reason = f"HTTP {he.code}"
        return jsonify(error="batch_failed", detail=f"Script-Ecom: {reason}"), 424
    except Exception as e:
        return jsonify(error="upstream_unreachable", detail=str(e)), 424

    statuses = load_order_print_statuses()
    printed = 0
    for o in payload.get("orders") or []:
        key = f"{o.get('platform')}|{o.get('order_id')}"
        st = statuses.get(key) or {}
        o["print_status"] = st.get("status") or "new"
        o["printed_at"] = st.get("printed_at") or ""
        if o["print_status"] == "printed":
            printed += 1
    payload["printed_orders"] = printed
    return jsonify(status="ok", batch=payload)


@app.route("/api/online/order-print-status", methods=["POST"])
@require_admin
def online_order_print_status():
    """ติ๊ก 'พิมพ์แล้ว' ระดับออเดอร์ (bulk) — ผูก order_id ดึงไฟล์ใหม่ทับกี่ครั้งสถานะก็ไม่หาย."""
    data = request.get_json(silent=True) or {}
    platform = str(data.get("platform", "")).strip().lower()
    action = str(data.get("action", "printed")).strip()
    batch_id = str(data.get("batch_id", "")).strip()
    order_ids = [str(x).strip() for x in (data.get("order_ids") or []) if str(x).strip()]
    if platform not in PRINT_PLATFORMS:
        return jsonify(error="bad_request", detail="platform ต้องเป็น shopee/lazada/tiktok"), 400
    if not order_ids:
        return jsonify(error="bad_request", detail="ต้องระบุ order_ids"), 400
    if action not in ("printed", "new"):
        return jsonify(error="bad_request", detail="action ต้องเป็น printed หรือ new"), 400
    statuses = load_order_print_statuses()
    now = datetime.now().isoformat(timespec="seconds")
    for oid in order_ids:
        key = f"{platform}|{oid}"
        if action == "printed":
            statuses[key] = {"status": "printed", "printed_at": now, "batch_id": batch_id}
        else:
            statuses.pop(key, None)
    save_order_print_statuses(statuses)
    return jsonify(status="ok", updated=len(order_ids), action=action)


@app.route("/api/online/print-files")
@require_admin
def online_print_files():
    date_iso = str(request.args.get("date", "")).strip()
    try:
        day = date.fromisoformat(date_iso).strftime("%d-%m-%Y")
    except ValueError:
        return jsonify(error="bad_request", detail="date must be YYYY-MM-DD"), 400

    groups = []
    total = 0
    statuses = load_print_statuses()
    for platform in PRINT_PLATFORMS:
        pdir = script_ecom_print_dir(date_iso, platform)
        files = []
        if pdir.exists():
            for path in sorted(pdir.glob("*.pdf"), key=lambda p: p.name.lower()):
                try:
                    size = path.stat().st_size
                except OSError:
                    size = 0
                meta = parse_print_filename(path.name, platform, day)
                key = print_status_key(date_iso, platform, path.name)
                status = statuses.get(key) or {}
                files.append({
                    "file": path.name,
                    "label": path.stem,
                    "size": size,
                    "orders": pdf_page_count(path),
                    "print_status": status.get("status") or "new",
                    "opened_at": status.get("opened_at") or "",
                    "printed_at": status.get("printed_at") or "",
                    **meta,
                })
        total += len(files)
        groups.append({"platform": platform, "count": len(files), "files": files})
    return jsonify(status="ok", date=date_iso, folder_date=day, total=total, groups=groups)


@app.route("/api/online/print-status", methods=["POST"])
@require_admin
def online_print_status():
    data = request.get_json(silent=True) or {}
    date_iso = str(data.get("date", "")).strip()
    platform = str(data.get("platform", "")).strip().lower()
    filename = os.path.basename(str(data.get("file", "")).strip())
    action = str(data.get("action", "")).strip().lower()
    if action not in {"opened", "printed", "reset"}:
        return jsonify(error="bad_request", detail="invalid action"), 400
    if not filename or not filename.lower().endswith(".pdf"):
        return jsonify(error="bad_request", detail="file must be a PDF filename"), 400
    try:
        pdir = script_ecom_print_dir(date_iso, platform).resolve()
    except ValueError:
        return jsonify(error="bad_request", detail="invalid date or platform"), 400
    path = (pdir / filename).resolve()
    if path.parent != pdir or not path.exists():
        return jsonify(error="not_found", detail="print file not found"), 404

    statuses = load_print_statuses()
    key = print_status_key(date_iso, platform, filename)
    now = datetime.now().isoformat(timespec="seconds")
    current = statuses.get(key) or {}
    if action == "reset":
        statuses.pop(key, None)
        result = {"status": "new", "opened_at": "", "printed_at": ""}
    else:
        if action == "opened" and current.get("status") != "printed":
            current["status"] = "opened"
            current["opened_at"] = current.get("opened_at") or now
        if action == "printed":
            current["status"] = "printed"
            current["opened_at"] = current.get("opened_at") or now
            current["printed_at"] = now
        current["updated_at"] = now
        statuses[key] = current
        result = {
            "status": current.get("status") or "new",
            "opened_at": current.get("opened_at") or "",
            "printed_at": current.get("printed_at") or "",
        }
    save_print_statuses(statuses)
    return jsonify(ok=True, **result)


@app.route("/api/online/print-file")
@require_admin
def online_print_file():
    date_iso = str(request.args.get("date", "")).strip()
    platform = str(request.args.get("platform", "")).strip().lower()
    filename = os.path.basename(str(request.args.get("file", "")).strip())
    if not filename or not filename.lower().endswith(".pdf"):
        return jsonify(error="bad_request", detail="file must be a PDF filename"), 400
    try:
        pdir = script_ecom_print_dir(date_iso, platform).resolve()
    except ValueError:
        return jsonify(error="bad_request", detail="invalid date or platform"), 400
    path = (pdir / filename).resolve()
    if path.parent != pdir:
        return jsonify(error="bad_request", detail="invalid file path"), 400
    if not path.exists():
        return jsonify(error="not_found", detail="print file not found"), 404
    return send_file(path, mimetype="application/pdf", as_attachment=False, download_name=filename)


if __name__ == "__main__":
    app.run(
        host=config.SERVER_HOST,
        port=config.SERVER_PORT,
        debug=os.getenv("FLASK_DEBUG") == "1",
        threaded=True,
    )
